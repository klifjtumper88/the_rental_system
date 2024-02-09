import openpyxl
def load_data():
    try:
        workbook = openpyxl.load_workbook("newrental.xlsx")
        sheet = workbook.active
        total_cars_rented = 0
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index == 1:  # Skip the header row
                continue
            total_cars_rented += int(row[1])  # Convert row[1] to an integer before adding
        return len(sheet['A']) - 1, total_cars_rented
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Renter Name", "Number of cars", "Rental basis", "Contact Details"])
        workbook.save('newrental.xlsx')
        return 100, 0
def rental_options(total_available_cars, cars_rented, sheet):
    option_basis = {1: "Hourly", 2: "Daily", 3: "Weekly", 4: "Monthly"}

    while True:
        options = input("""
Choose within the provided options using numbers:
1. A rental car for hourly basis
2. A rental car for daily basis
3. A rental car for weekly basis
4. A rental car for monthly basis
5. Current sheet status
6. Exit
Enter your choice: """)

        try:
            options = int(options)
        except ValueError:
            print("Invalid choice. Please enter a number between 1 and 6.")
            continue

        if options in [1, 2, 3, 4]:
            rental_basis = option_basis[options]
            print(f"You chose rental car for {rental_basis} basis")
            req_cars = int(input("How many cars do you want: "))
            if req_cars <= total_available_cars - cars_rented:
                if req_cars + cars_rented <= total_available_cars:
                    renter_name = input("Enter your name: ").upper()
                    renter_phone_no = input("Enter contact number: ")
                    total_available_cars -= req_cars
                    cars_rented += req_cars
                    print(f"You have successfully rented {req_cars} car(s).")
                    print(
                        f"So you will be charged {'$10 per hour' if options == 1 else '$30 per day' if options == 2 else '$50 per week' if options == 3 else '$150 per month'} per car.")
                    sheet.append([renter_name, req_cars, rental_basis, renter_phone_no])
                    print(f"Total Cars Rented: {cars_rented}")
                    print(f"Total Cars Available: {total_available_cars}")
                    break
                else:
                    print("Sorry, no more cars available for rent.")
            else:
                print("Sorry, requested number of cars not available.")
        elif options == 5:
            print("Current Rental Status:")
            for row in sheet.iter_rows(values_only=True):
                print(row)
            print(f"Total Cars Rented: {cars_rented}")
            print(f"Total Cars Available: {total_available_cars}")
        elif options == 6:
            print("Exiting program. Goodbye!")
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 6.")

total_available_cars, cars_rented = load_data()
workbook = openpyxl.load_workbook("newrental.xlsx")
sheet = workbook.active
rental_options(total_available_cars, cars_rented, sheet)
